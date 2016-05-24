import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
import calendar as cl
import csv
print("Let's get started.")

wb = openpyxl.load_workbook('Corn Futures values.xlsx')
sheet = wb.get_sheet_by_name('Corn Futures')
print("Data loaded.")

row = len(sheet.columns[0])
col = len(sheet.rows[0])

cornData = []
cornData.append(["Date", "Contract", "Price"])

for i in range(0, col, 3):
    year = sheet.cell(row=0, column=i).value
    monthInt = sheet.cell(row=2, column=i).value
    month = cl.month_abbr[monthInt]
    contract = month + str(year)
    for j in range(7, row):
        item = []
        date = sheet.cell(row=j, column=i).value
        price = sheet.cell(row=j, column=i+1).value
        if date:
            item.append(date)
            item.append(contract)
            item.append(price)
            cornData.append(item)
        else:
            break
print("List created successfully. There're " + str(len(cornData)) + " items in the lsit.")

csvfile = "CornFuturesData.csv"
with open(csvfile, "w") as output:
    writer = csv.writer(output, lineterminator = '\n')
    writer.writerows(cornData)
print("Data exported in csv format.")

txtfile = open('CornFuturesData.txt','w')
for item in cornData:
    txtfile.write(str(item) +'\n')
print("Data exported in txt format.")
